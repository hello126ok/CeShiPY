from openpyxl import load_workbook
filename = '区县代码处理中间表.xlsx'
wb = load_workbook(filename)

print(f"{filename}里面的工作表是{wb.sheetnames}")

ws = wb['Sheet3']
ws.sheet_properties.tabColor = '808080'

# 读取第二行作为列名
header_row_num = 2
header_row = ws[header_row_num]  # 索引从1开始，第二行为索引2

headers = [cell.value for cell in header_row if cell.value is not None]
column_map = {name: idx+1 for idx, name in enumerate(headers)}  # 列索引从1开始
print(column_map)

''' #下面代码和上面的三行代码等价
column_map = {}
for idx, cell in enumerate(header_row):
        if cell.value is not None:
            column_map[cell.value] = idx + 1  # 列索引从1开始
print(column_map)
'''
'''
# 读取数据
data = []
for row in ws.iter_rows(min_row=header_row_num+1, values_only=True):
    row_data = {}
    for col_name in column_map:
        col_idx = column_map[col_name]
        row_data[col_name] = row[col_idx-1]  # 转换为0-based索引
    data.append(row_data)
print(data)
'''
# 遍历所有行（包含空单元格）
for row in ws.iter_rows():
    for cell in row:
        a = 1
       # cell.value = None if cell.value is not None else cell.value
       # print(cell.coordinate, cell.value, type(cell.value))

'''
# 插入行和列
ws.insert_cols(idx=3,amount=2)
ws.insert_rows(idx=3,amount=4)
ws.title = "我是修改后的sheet名"
'''
print(ws.max_row)
print(ws.max_column)

wb.save(filename)
wb.close()

